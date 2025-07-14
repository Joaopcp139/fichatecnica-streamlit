import streamlit as st
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from io import BytesIO
import re

st.set_page_config(page_title="Ficha Técnica Automática", layout="centered")
st.title("📄 Gerador de Ficha Técnica a partir da OP")

# Upload do PDF da OP
uploaded_pdf = st.file_uploader("📎 Envie o arquivo da OP (PDF)", type="pdf")

# Seleção do tipo de ficha
ficha_tipo = st.radio("Tipo de ficha técnica:", ["SACO", "FILME"])

# Função para extrair informações do PDF
def extrair_dados(texto):
    def extrair(padrao, default=""):
        match = re.search(padrao, texto)
        return match.group(1).strip() if match else default

    return {
        "cliente": extrair(r"Cliente:\s*(.+)"),
        "produto": extrair(r"Produto:\s*(.+)"),
        "codigo_produto": extrair(r"(\d{5,})\s*-\s*"),
        "data_pedido": extrair(r"Data do Pedido:\s*(\d{2}/\d{2}/\d{4})"),
        "data_entrega": extrair(r"Data de Entrega:\s*(\d{2}/\d{2}/\d{4})"),
        "pedido_numero": extrair(r"Pedido Nº:\s*(\d+)"),
        "largura": extrair(r"Largura:\s*(\d+)"),
        "espessura": extrair(r"Espessura:\s*([0-9,\.]+)"),
        "passo": extrair(r"Passo:\s*(\d+)"),
        "cilindro": extrair(r"Cilindro:\s*(\d+)"),
        "quantidade_kg": extrair(r"Quantidade \(KG\):\s*([0-9\.]+)"),
        "quantidade_bobinas": extrair(r"Quantidade de bobinas:\s*(\d+)"),
        "tubete": "Yes" if "Tubete 3: Sim" in texto else "No",
        "laminado": "Sim" if "Laminado: Sim" in texto else "Não",
        "sanfona": "Sim" if "Sanfona Sim" in texto else "Não",
        "materia_prima": "Yes" if "Matéria-prima PE: Sim" in texto else "No",
        "frente1": "Yes" if "Frente 1: Yes" in texto else "No",
        "oc": extrair(r"OC:\s*(\d+)")
    }

# Quando o botão for clicado
if uploaded_pdf and st.button("🔄 Gerar ficha técnica"):
    try:
        # Leitura do PDF
        pdf_reader = PdfReader(uploaded_pdf)
        texto = "\n".join([page.extract_text() for page in pdf_reader.pages])
        
        # Extração dos dados
        dados = extrair_dados(texto)

        # Escolha do modelo correto
        modelo_arquivo = "FILME.xlsx" if ficha_tipo == "FILME" else "SACO.xlsx"
        wb = load_workbook(modelo_arquivo)
        ws = wb.active

        # Preenchimento automático básico (exemplo para alguns campos)
        try:
            ws["D6"] = dados["cliente"]
            ws["F6"] = dados["codigo_produto"]
            ws["D7"] = dados["produto"]
            ws["B13"] = dados["largura"]
            ws["D13"] = dados["passo"]
            ws["F13"] = dados["espessura"]
        except:
            pass  # Ignora se a célula não existir ou der erro

        # Exporta o resultado
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        st.success("✅ Ficha técnica gerada com sucesso!")
        st.download_button(
            label="📥 Baixar ficha preenchida",
            data=output,
            file_name="ficha_tecnica_preenchida.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"❌ Erro ao processar a OP: {e}")
        # Preenchimento automático básico (exemplo para alguns campos)
try:
    ws["D6"] = dados["cliente"]
    ws["F6"] = dados["codigo_produto"]
    ws["D7"] = dados["produto"]
    ws["B13"] = dados["largura"]
    ws["D13"] = dados["passo"]
    ws["F13"] = dados["espessura"]
except:
    pass  # Ignora se a célula não existir ou der erro
por isso aqui:
python
Copiar código
from datetime import datetime

# Preenchimento automático básico (exemplo para alguns campos)
try:
    ws["D6"] = dados["cliente"]
    ws["F6"] = dados["codigo_produto"]
    ws["D7"] = dados["produto"]
    ws["B13"] = dados["largura"]
    ws["D13"] = dados["passo"]
    ws["F13"] = dados["espessura"]

    # 🗓️ Insere a data do dia na célula L2
    hoje = datetime.today().strftime("%d/%m/%Y")
    ws["L2"] = hoje
except:
    pass  # Ignora se der erro
